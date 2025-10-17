from __future__ import annotations

import logging
import calendar
from datetime import datetime, timedelta, date
from collections import defaultdict
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from typing import Dict, List, Optional
from django.db import ProgrammingError
from django.db.models import Sum, Count, Avg
from django.db.models.functions import TruncDate
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import ValidationError, NotFound
from drf_spectacular.utils import extend_schema, OpenApiParameter
from django.http import HttpResponse
import csv

from apps.orders.models import ProductOrder
from apps.payments.models import Deposit
from apps.payouts.models import Payout
from apps.users.models import User
from apps.providers.models import Integration
from apps.payments.models import PaymentMethod
from apps.providers.adapters import resolve_adapter_credentials
from apps.currencies.models import Currency
from .models import CapitalAdjustment
from .serializers import (
    OverviewResponseSerializer,
    DailyResponseSerializer,
    CapitalAdjustmentSerializer,
    CapitalAdjustmentInputSerializer,
)
from apps.users.permissions import RequireAdminRole
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from django.db.models import Q


logger = logging.getLogger(__name__)


def _resolve_tenant_id(request) -> str | None:
    tid = getattr(request, 'tenant', None)
    if tid and getattr(tid, 'id', None):
        return str(tid.id)
    user = getattr(request, 'user', None)
    if user and getattr(user, 'tenant_id', None):
        return str(user.tenant_id)
    return None


def _parse_dates(request):
    # Accept ISO date strings; default last 30 days
    q_from = (request.query_params.get('from') or '').strip()
    q_to = (request.query_params.get('to') or '').strip()
    try:
        dfrom = datetime.fromisoformat(q_from).date() if q_from else (date.today() - timedelta(days=29))
    except Exception:
        dfrom = date.today() - timedelta(days=29)
    try:
        dto = datetime.fromisoformat(q_to).date() if q_to else date.today()
    except Exception:
        dto = date.today()
    if dfrom > dto:
        dfrom, dto = dto, dfrom
    # Inclusive end
    return dfrom, dto


def _money_totals(rows, currency_field: str, amount_field: str):
    sums = defaultdict(lambda: 0)
    for r in rows:
        cur = getattr(r, currency_field) or 'USD'
        amt = getattr(r, amount_field) or 0
        try:
            sums[cur] += float(amt)
        except Exception:
            pass
    return [{ 'currency': k, 'amount': round(v, 6) } for k, v in sorted(sums.items())]


def _normalize_currency_code(raw: str | None) -> str:
    if not raw:
        return 'USD'
    return raw.strip().upper() or 'USD'


def _decimal_from(value) -> Decimal:
    if isinstance(value, Decimal):
        return value
    if value is None:
        return Decimal('0')
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal('0')


def _to_float(value: Decimal, places: int = 4) -> float:
    if not isinstance(value, Decimal):
        value = _decimal_from(value)
    quant = Decimal('1').scaleb(-places)
    return float(value.quantize(quant, rounding=ROUND_HALF_UP))


def _infer_provider_currency(integration: Integration) -> str:
    possible = getattr(integration, 'balance_currency', None)
    if isinstance(possible, str) and possible.strip():
        return possible.strip().upper()
    raw_settings = getattr(integration, 'settings', None)
    settings = raw_settings if isinstance(raw_settings, dict) else None
    if settings:
        for key in ('balanceCurrency', 'balance_currency', 'currency', 'currencyCode', 'code'):
            val = settings.get(key)
            if isinstance(val, str) and val.strip():
                return val.strip().upper()
    provider_code = (integration.provider or '').lower()
    name_code = (integration.name or '').lower()
    defaults = {
        'internal': 'USD',
        'barakat': 'TRY',
        'znet': 'TRY',
        'apstore': 'TRY',
        'shamtech': 'TRY',
    }
    if name_code in defaults:
        return defaults[name_code]
    return defaults.get(provider_code, 'USD')


def _normalize_label_value(label: Optional[str]) -> str:
    if label is None:
        return '—'
    value = str(label).strip()
    return value or '—'


def _parse_decimal_or_none(value) -> Optional[Decimal]:
    if value is None:
        return None
    try:
        raw = str(value).strip()
    except Exception:
        return None
    if not raw:
        return None
    normalized = raw.replace(',', '.')
    try:
        return Decimal(normalized)
    except (InvalidOperation, TypeError, ValueError):
        return None


def _quantize_three_places(value: Decimal) -> Decimal:
    return value.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)


def _refresh_znet_live_values(integration: Integration) -> Optional[str]:
    binding, creds = resolve_adapter_credentials(
        integration.provider,
        base_url=integration.base_url,
        api_token=getattr(integration, 'api_token', None),
        kod=getattr(integration, 'kod', None),
        sifre=getattr(integration, 'sifre', None),
    )
    if not binding or not creds:
        return None

    try:
        result = binding.adapter.get_balance(creds) or {}
    except Exception as exc:  # pragma: no cover - defensive fetch guard
        logger.warning(
            'Failed to fetch ZNET balance/debt',
            extra={'integration_id': str(integration.id), 'reason': str(exc)[:120]},
        )
        return None

    if not isinstance(result, dict):
        return None
    if result.get('error') or result.get('missingConfig'):
        logger.warning(
            'ZNET adapter reported error while fetching debt',
            extra={
                'integration_id': str(integration.id),
                'error': result.get('error'),
                'message': str(result.get('message') or '')[:120],
            },
        )
        return None

    balance_val = _parse_decimal_or_none(result.get('balance'))
    debt_val = _parse_decimal_or_none(result.get('debt'))

    updates: dict = {}
    now = timezone.now()
    if balance_val is not None:
        updates['balance'] = _quantize_three_places(balance_val)
        updates['balance_updated_at'] = now
    if debt_val is not None:
        updates['debt'] = _quantize_three_places(debt_val)
        updates['debt_updated_at'] = now

    if updates:
        Integration.objects.filter(id=integration.id).update(**updates)
        integration.balance = updates.get('balance', integration.balance)
        integration.debt = updates.get('debt', integration.debt)
        integration.balance_updated_at = updates.get('balance_updated_at', integration.balance_updated_at)
        integration.debt_updated_at = updates.get('debt_updated_at', integration.debt_updated_at)
        balance_currency = getattr(integration, 'balance_currency', None)
        debt_currency = getattr(integration, 'debt_currency', None)
        fallback_currency = getattr(integration, 'currency', None)
        return balance_currency or debt_currency or fallback_currency

    return None


def _parse_iso_date(raw: Optional[str]) -> Optional[date]:
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw).date()
    except Exception:
        return None


def _shift_month(base: date, offset: int) -> date:
    month_index = base.month - 1 + offset
    year = base.year + month_index // 12
    month = month_index % 12 + 1
    last_day = calendar.monthrange(year, month)[1]
    day = min(base.day, last_day)
    return date(year, month, day)


def _resolve_date_range(range_key: Optional[str], start_str: Optional[str], end_str: Optional[str]) -> tuple[date, date, str, bool]:
    today = timezone.localdate()
    key = (range_key or 'today').lower()

    if key == 'today':
        return today, today, 'today', True

    if key == 'this_month':
        start_date = today.replace(day=1)
        return start_date, today, 'this_month', True

    if key == 'last_month':
        first_this_month = today.replace(day=1)
        previous_month_last = first_this_month - timedelta(days=1)
        return previous_month_last.replace(day=1), previous_month_last, 'last_month', True

    if key == 'last_6_months':
        start_date = _shift_month(today.replace(day=1), -5)
        return start_date, today, 'last_6_months', True

    if key == 'custom':
        start_date = _parse_iso_date(start_str)
        end_date = _parse_iso_date(end_str)
        if not start_date or not end_date:
            return today, today, 'custom', False
        if start_date > end_date:
            start_date, end_date = end_date, start_date
        return start_date, end_date, 'custom', True

    # Fallback: treat invalid range as today
    return today, today, 'today', True

def _build_capital_workbook(summary: dict) -> Workbook:
    wb = Workbook()

    header_fill = PatternFill(fill_type='solid', fgColor='2563EB')
    header_font = Font(color='FFFFFF', bold=True)
    header_align = Alignment(horizontal='center', vertical='center')
    section_fill = PatternFill(fill_type='solid', fgColor='1E3A8A')
    section_font = Font(color='FFFFFF', bold=True, size=12)
    body_align_right = Alignment(horizontal='right', vertical='center')
    body_align_center = Alignment(horizontal='center', vertical='center')
    body_align_left = Alignment(horizontal='left', vertical='center')
    thin_side = Side(style='thin', color='D1D5DB')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def _set_column_width(ws, col_idx: int, width: int) -> None:
        column_letter = get_column_letter(col_idx)
        current_width = ws.column_dimensions[column_letter].width
        if current_width is None or current_width < width:
            ws.column_dimensions[column_letter].width = width

    def _render_table(
        ws,
        start_row: int,
        headers: list[str],
        widths: list[int],
        rows: list[list],
        number_cols: set[int],
        center_cols: set[int] | None = None,
        footer: list | None = None,
        footer_number_cols: set[int] | None = None,
        footer_bold: bool = True,
    ) -> int:
        for idx, (header, width) in enumerate(zip(headers, widths), start=1):
            cell = ws.cell(row=start_row, column=idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            _set_column_width(ws, idx, width)

        data_start = start_row + 1
        for row_idx, row_values in enumerate(rows, start=data_start):
            for col_idx, value in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx in number_cols and value is not None:
                    cell.number_format = '#,##0.00'
                    cell.alignment = body_align_right
                elif center_cols and col_idx in center_cols:
                    cell.alignment = body_align_center
                else:
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.alignment = body_align_right
                    else:
                        cell.alignment = body_align_left
                cell.border = thin_border

        next_row = start_row + 1 + len(rows)

        if footer is not None:
            footer_number_cols = footer_number_cols or number_cols
            for col_idx, value in enumerate(footer, start=1):
                cell_value = value if value is not None else ''
                cell = ws.cell(row=next_row, column=col_idx, value=cell_value)
                cell.border = thin_border
                if col_idx in footer_number_cols and value not in (None, ''):
                    cell.number_format = '#,##0.00'
                    cell.alignment = body_align_right
                elif isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = body_align_right
                else:
                    if col_idx == 1:
                        cell.alignment = body_align_left
                    else:
                        cell.alignment = body_align_center
                if footer_bold:
                    cell.font = Font(bold=True)
            next_row += 1

        return next_row

    def _render_section_title(ws, row: int, title: str, column_count: int) -> None:
        for col_idx in range(1, column_count + 1):
            cell = ws.cell(row=row, column=col_idx)
            if col_idx == 1:
                cell.value = title
                cell.alignment = body_align_left
            else:
                cell.value = ''
                cell.alignment = body_align_center
            cell.font = section_font
            cell.fill = section_fill
            cell.border = thin_border

    def _render_section(
        ws,
        start_row: int,
        title: str,
        headers: list[str],
        widths: list[int],
        rows: list[list],
        number_cols: set[int],
        center_cols: set[int] | None = None,
        footer: list | None = None,
        footer_number_cols: set[int] | None = None,
    ) -> int:
        _render_section_title(ws, start_row, title, len(headers))
        next_row = _render_table(
            ws,
            start_row + 1,
            headers,
            widths,
            rows,
            number_cols,
            center_cols=center_cols,
            footer=footer,
            footer_number_cols=footer_number_cols,
        )
        return next_row + 1

    def _safe_parse_iso(value: Optional[str]) -> Optional[str]:
        if not value:
            return None
        try:
            dt = datetime.fromisoformat(value)
            return dt.strftime('%Y-%m-%d %H:%M')
        except Exception:
            return value

    sorted_currencies: list[str] = summary.get('sortedCurrencies') or []
    if not sorted_currencies:
        currency_universe: set[str] = set()
        for section in ('providers', 'users', 'adjustments'):
            items = summary.get(section, {}).get('items', []) if section != 'users' else []
            if section == 'users':
                items = summary.get('users', {}).get('totals', [])
            for item in items:
                currency = _normalize_currency_code(item.get('currency'))
                currency_universe.add(currency)
        sorted_currencies = sorted(currency_universe)

    payment_method_labels: list[str] = summary.get('paymentMethods') or []
    payment_method_set = set(payment_method_labels)

    adjustment_items = summary.get('adjustments', {}).get('items', [])
    adjustment_matrix: Dict[str, Dict[str, Decimal]] = defaultdict(lambda: defaultdict(lambda: Decimal('0')))
    adjustment_notes: Dict[str, str] = {}
    for item in adjustment_items:
        label = _normalize_label_value(item.get('label'))
        currency = _normalize_currency_code(item.get('currency'))
        amount = _decimal_from(item.get('amount'))
        adjustment_matrix[label][currency] += amount
        note = (item.get('note') or '').strip()
        if note:
            adjustment_notes[label] = note

    manual_labels = sorted(label for label in adjustment_matrix.keys() if label not in payment_method_set)

    ws_report = wb.active
    ws_report.title = 'تقرير رأس المال'
    current_row = 1

    provider_items = summary.get('providers', {}).get('items', [])
    summary_rows = [
        ('إجمالي رأس المال (USD)', summary.get('grandTotalUsd', 0.0)),
        ('إجمالي أرصدة المزودين (USD)', summary.get('providers', {}).get('totalUsd', 0.0)),
        ('إجمالي أرصدة المستخدمين (USD)', summary.get('users', {}).get('totalUsd', 0.0)),
        ('إجمالي التعديلات (USD)', summary.get('adjustments', {}).get('totalUsd', 0.0)),
        ('عدد المزودين', len(provider_items)),
        ('وسائل الدفع الفعّالة', len(payment_method_labels)),
        ('عدد المستخدمين', summary.get('users', {}).get('count', 0)),
        ('عدد التعديلات المسجلة', len(adjustment_items)),
        ('عملات بدون سعر تحويل', ', '.join(summary.get('missingRates', [])) or '-'),
    ]

    current_row = _render_section(
        ws_report,
        current_row,
        'القيم الإجمالية',
        headers=['البند', 'القيمة'],
        widths=[38, 28],
        rows=summary_rows,
        number_cols={2},
    )

    provider_rows: list[list] = []
    for item in provider_items:
        provider_rows.append([
            item.get('name') or '',
            item.get('provider') or '-',
            item.get('balance'),
            item.get('debt'),
            item.get('netBalance'),
            item.get('balanceUsd'),
            item.get('currency'),
            _safe_parse_iso(item.get('balanceUpdatedAt')),
            _safe_parse_iso(item.get('debtUpdatedAt')),
        ])

    provider_footer = ['الإجمالي (USD)', '', '', '', '', summary.get('providers', {}).get('totalUsd', 0.0), '', '', '']
    current_row = _render_section(
        ws_report,
        current_row,
        'تفاصيل المزودين',
        headers=['المزود', 'النوع', 'الرصيد', 'الدين', 'المحصلة', 'المحصلة (USD)', 'العملة', 'آخر تحديث للرصيد', 'آخر تحديث للدين'],
        widths=[28, 16, 16, 16, 16, 18, 10, 24, 24],
        rows=provider_rows,
        number_cols={3, 4, 5, 6},
        center_cols={2, 7, 8, 9},
        footer=provider_footer,
        footer_number_cols={6},
    )

    user_rows: list[list] = []
    for row in summary.get('users', {}).get('totals', []):
        user_rows.append([
            row.get('currency'),
            row.get('amount'),
            row.get('amountUsd'),
        ])

    user_footer = ['الإجمالي (USD)', '', summary.get('users', {}).get('totalUsd', 0.0)]
    current_row = _render_section(
        ws_report,
        current_row,
        'أرصدة المستخدمين',
        headers=['العملة', 'الرصيد', 'بالدولار'],
        widths=[12, 18, 18],
        rows=user_rows,
        number_cols={2, 3},
        center_cols={1},
        footer=user_footer,
        footer_number_cols={3},
    )

    payment_headers = ['الجهة'] + sorted_currencies + ['ملاحظة']
    payment_widths = [28] + [16 for _ in sorted_currencies] + [32]
    payment_rows: list[list] = []
    for label in payment_method_labels:
        row_values: list = [label]
        for currency in sorted_currencies:
            amount = adjustment_matrix.get(label, {}).get(currency)
            row_values.append(float(amount) if amount is not None else 0.0)
        row_values.append(adjustment_notes.get(label, ''))
        payment_rows.append(row_values)

    current_row = _render_section(
        ws_report,
        current_row,
        'وسائل الدفع الفعّالة',
        headers=payment_headers,
        widths=payment_widths,
        rows=payment_rows,
        number_cols=set(range(2, 2 + len(sorted_currencies))),
        center_cols=set(range(2, 2 + len(sorted_currencies))),
    )

    manual_headers = ['الجهة'] + sorted_currencies + ['ملاحظة']
    manual_widths = [28] + [16 for _ in sorted_currencies] + [32]
    manual_rows: list[list] = []
    for label in manual_labels:
        row_values: list = [label]
        for currency in sorted_currencies:
            amount = adjustment_matrix.get(label, {}).get(currency)
            row_values.append(float(amount) if amount is not None else 0.0)
        row_values.append(adjustment_notes.get(label, ''))
        manual_rows.append(row_values)

    current_row = _render_section(
        ws_report,
        current_row,
        'صفوف يدوية',
        headers=manual_headers,
        widths=manual_widths,
        rows=manual_rows,
        number_cols=set(range(2, 2 + len(sorted_currencies))),
        center_cols=set(range(2, 2 + len(sorted_currencies))),
    )

    adjustment_rows: list[list] = []
    for row in adjustment_items:
        adjustment_rows.append([
            _normalize_label_value(row.get('label')),
            row.get('currency'),
            row.get('amount'),
            row.get('amountUsd'),
            row.get('note') or '',
            _safe_parse_iso(row.get('createdAt')),
            _safe_parse_iso(row.get('updatedAt')),
        ])

    adjustments_footer = ['الإجمالي (USD)', '', '', summary.get('adjustments', {}).get('totalUsd', 0.0), '', '', '']
    current_row = _render_section(
        ws_report,
        current_row,
        'التعديلات التفصيلية',
        headers=['العنوان', 'العملة', 'القيمة', 'بالدولار', 'ملاحظة', 'أنشئت في', 'آخر تعديل'],
        widths=[28, 10, 16, 16, 28, 22, 22],
        rows=adjustment_rows,
        number_cols={3, 4},
        center_cols={2, 6, 7},
        footer=adjustments_footer,
        footer_number_cols={4},
    )

    totals_rows_data = summary.get('overallTotals', [])
    totals_rows = [[row.get('currency'), row.get('amount')] for row in totals_rows_data]
    totals_footer = ['USD (المحصلة)', summary.get('grandTotalUsd', 0.0)]
    current_row = _render_section(
        ws_report,
        current_row,
        'إجمالي الأرصدة حسب العملة',
        headers=['العملة', 'الإجمالي'],
        widths=[12, 18],
        rows=totals_rows,
        number_cols={2},
        center_cols={1},
        footer=totals_footer,
        footer_number_cols={2},
    )

    rate_rows = sorted(summary.get('rates', {}).items(), key=lambda item: item[0])
    current_row = _render_section(
        ws_report,
        current_row,
        'أسعار الصرف الحالية',
        headers=['العملة', 'السعر مقابل USD'],
        widths=[12, 22],
        rows=[[code, rate] for code, rate in rate_rows],
        number_cols={2},
        center_cols={1},
    )

    return wb


class AdminReportsUsersView(APIView):
    permission_classes = [IsAuthenticated, RequireAdminRole]

    @extend_schema(tags=["Admin Reports"], responses={200: None})
    def get(self, request):
        tenant_id = _resolve_tenant_id(request)
        if not tenant_id:
            raise ValidationError('TENANT_ID_REQUIRED')

        from apps.orders.models import LegacyUser

        q = (request.query_params.get('q') or '').strip()
        limit = int(request.query_params.get('limit') or 20)
        limit = max(1, min(limit, 100))

        qs = LegacyUser.objects.filter(tenant_id=tenant_id)

        if q:
            qs = qs.filter(Q(email__icontains=q) | Q(username__icontains=q))

        items = list(qs.only('id', 'email', 'username').order_by('-id')[:limit])

        def label(user_obj):
            username = getattr(user_obj, 'username', None)
            email = getattr(user_obj, 'email', None)
            return username or email or str(user_obj.id)

        return Response([{'id': str(u.id), 'label': label(u)} for u in items])


class AdminReportsProvidersView(APIView):
    permission_classes = [IsAuthenticated, RequireAdminRole]

    @extend_schema(tags=["Admin Reports"], responses={200: None})
    def get(self, request):
        tenant_id = _resolve_tenant_id(request)
        if not tenant_id:
            raise ValidationError('TENANT_ID_REQUIRED')
        
        # جلب المزودين من جدول integrations
        from apps.providers.models import Integration
        
        providers = []
        
        # إضافة "يدوي" كخيار أول
        providers.append({ 'id': 'manual', 'label': 'يدوي' })
        
        # جلب المزودين من جدول integrations
        integrations = Integration.objects.filter(
            tenant_id=tenant_id,
            enabled=True
        ).only('id', 'name', 'provider').order_by('name')
        
        for integration in integrations:
            # استخدام الاسم إن وُجد، وإلا نوع المزود
            label = integration.name or integration.provider or str(integration.id)
            providers.append({ 
                'id': str(integration.id), 
                'label': label
            })
        
        return Response(providers)


class AdminReportsProfitsView(APIView):
    permission_classes = [IsAuthenticated, RequireAdminRole]
    _REJECTED_STATUSES = ('rejected', 'cancelled', 'canceled')

    @extend_schema(tags=["Admin Reports"], responses={200: None})
    def get(self, request):
        tenant_id = _resolve_tenant_id(request)
        if not tenant_id:
            raise ValidationError('TENANT_ID_REQUIRED')

        range_param = (request.query_params.get('range') or 'today').strip().lower()
        start_param = (request.query_params.get('start') or '').strip() or None
        end_param = (request.query_params.get('end') or '').strip() or None

        start_date, end_date, resolved_range, valid_range = _resolve_date_range(range_param, start_param, end_param)
        if range_param == 'custom' and not valid_range:
            raise ValidationError('INVALID_DATE_RANGE')

        view_param = (request.query_params.get('view') or '').strip().lower()
        if view_param not in ('usd_only',):
            view_param = ''

        user_id = (request.query_params.get('userId') or '').strip()
        provider_param = (request.query_params.get('provider') or '').strip()

        filters_q = Q(tenant_id=tenant_id)
        if user_id:
            filters_q &= Q(user_id=user_id)

        if provider_param == 'manual':
            filters_q &= (Q(provider_id__isnull=True) | Q(provider_id__exact=''))
        elif provider_param:
            filters_q &= Q(provider_id=provider_param)

        approved_qs = ProductOrder.objects.filter(filters_q, approved_local_date__isnull=False)
        approved_qs = approved_qs.filter(approved_local_date__gte=start_date, approved_local_date__lte=end_date)

        rejected_qs = ProductOrder.objects.filter(filters_q, status__in=self._REJECTED_STATUSES)
        rejected_qs = rejected_qs.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)

        approved_count = approved_qs.count()
        rejected_count = rejected_qs.count()
        total_count = approved_count + rejected_count

        aggregates = approved_qs.aggregate(
            total_cost_try=Sum('cost_try_at_approval'),
            total_sales_try=Sum('sell_try_at_approval'),
            total_profit_try=Sum('profit_try_at_approval'),
            total_cost_usd=Sum('cost_usd_at_order'),
            total_sales_usd=Sum('sell_usd_at_order'),
            total_profit_usd=Sum('profit_usd_at_approval'),
            avg_rate=Avg('fx_usd_try_at_approval'),
        )

        def _dec(value) -> Decimal:
            if isinstance(value, Decimal):
                return value
            if value is None:
                return Decimal('0')
            try:
                return Decimal(str(value))
            except (InvalidOperation, TypeError, ValueError):
                return Decimal('0')

        total_cost_try = _dec(aggregates.get('total_cost_try'))
        total_sales_try = _dec(aggregates.get('total_sales_try'))
        profit_try = aggregates.get('total_profit_try')
        profit_try = _dec(profit_try) if profit_try is not None else (total_sales_try - total_cost_try)

        total_cost_usd = _dec(aggregates.get('total_cost_usd'))
        total_sales_usd = _dec(aggregates.get('total_sales_usd'))
        profit_usd = aggregates.get('total_profit_usd')
        profit_usd = _dec(profit_usd) if profit_usd is not None else (total_sales_usd - total_cost_usd)

        avg_rate = aggregates.get('avg_rate')
        if avg_rate is None and total_sales_usd and total_sales_usd != 0:
            try:
                avg_rate = total_sales_try / total_sales_usd
            except (InvalidOperation, ZeroDivisionError):
                avg_rate = None
        avg_rate = _dec(avg_rate)

        filters_payload = {
            'range': resolved_range,
            'start': start_date.isoformat(),
            'end': end_date.isoformat(),
            'userId': user_id or None,
            'provider': provider_param or None,
        }

        counts_payload = {
            'total': total_count,
            'approved': approved_count,
            'rejected': rejected_count,
        }

        if view_param == 'usd_only':
            filters_payload['view'] = 'usd_only'
            body = {
                'filters': filters_payload,
                'counts': counts_payload,
                'totalsUSD': {
                    'cost': _to_float(total_cost_usd, places=2),
                    'sales': _to_float(total_sales_usd, places=2),
                },
                'profitUSD': _to_float(profit_usd, places=2),
                'rateTRY': _to_float(avg_rate, places=6),
            }
            return Response(body)

        body = {
            'filters': filters_payload,
            'counts': counts_payload,
            'totalsTRY': {
                'cost': _to_float(total_cost_try, places=2),
                'sales': _to_float(total_sales_try, places=2),
            },
            'profit': {
                'try': _to_float(profit_try, places=2),
                'usd': _to_float(profit_usd, places=2),
                'rateTRY': _to_float(avg_rate, places=6),
            },
        }
        return Response(body)


def _build_capital_summary_payload(tenant_id: str) -> dict:
    rates_qs = Currency.objects.filter(tenant_id=tenant_id).order_by('code')
    rates_map: Dict[str, Decimal] = {}
    primary_currency: Optional[str] = None
    currency_universe: set[str] = set()

    for cur in rates_qs:
        code = _normalize_currency_code(getattr(cur, 'code', None))
        rate_value = getattr(cur, 'rate', None)
        try:
            rate = Decimal(str(rate_value)) if rate_value is not None else None
        except (InvalidOperation, TypeError, ValueError):
            rate = None
        if rate and rate > 0:
            rates_map[code] = rate
        currency_universe.add(code)
        if getattr(cur, 'is_primary', False) and not primary_currency:
            primary_currency = code

    rates_map.setdefault('USD', Decimal('1'))
    currency_universe.add('USD')

    missing_rates: set[str] = set()

    def convert_to_usd(amount: Decimal, currency_code: str | None) -> Decimal:
        code = _normalize_currency_code(currency_code)
        rate = rates_map.get(code)
        if not rate or rate <= 0:
            if code:
                missing_rates.add(code)
            return Decimal('0')
        if code == 'USD':
            return amount
        try:
            return amount / rate
        except (InvalidOperation, ZeroDivisionError):
            missing_rates.add(code)
            return Decimal('0')

    users_qs = (
        User.objects.filter(tenant_id=tenant_id)
        .exclude(status='disabled')
        .values('currency')
        .annotate(total=Sum('balance'), count=Count('id'))
    )

    users_totals: List[dict] = []
    users_total_usd = Decimal('0')
    users_count = 0
    overall_totals_map: Dict[str, Decimal] = defaultdict(lambda: Decimal('0'))

    for row in users_qs:
        currency = _normalize_currency_code(row.get('currency'))
        amount = _decimal_from(row.get('total'))
        usd_amount = convert_to_usd(amount, currency)
        users_total_usd += usd_amount
        users_count += int(row.get('count') or 0)
        users_totals.append({
            'currency': currency,
            'amount': _to_float(amount),
            'amountUsd': _to_float(usd_amount),
        })
        overall_totals_map[currency] += amount
        currency_universe.add(currency)

    providers_qs = Integration.objects.filter(tenant_id=tenant_id, enabled=True)

    providers_items: List[dict] = []
    providers_totals_map: Dict[str, Decimal] = defaultdict(lambda: Decimal('0'))
    providers_total_usd = Decimal('0')
    for integration in providers_qs:
        provider_key = (integration.provider or '').strip().lower()

        live_currency = None
        if provider_key == 'znet':
            live_currency = _refresh_znet_live_values(integration)

        amount = _decimal_from(getattr(integration, 'balance', None))
        debt = _decimal_from(getattr(integration, 'debt', None))
        currency = _normalize_currency_code(live_currency) if live_currency else _infer_provider_currency(integration)

        net_amount = amount - debt

        usd_amount = convert_to_usd(net_amount, currency)
        providers_total_usd += usd_amount
        providers_totals_map[currency] += net_amount
        overall_totals_map[currency] += net_amount
        currency_universe.add(currency)

        providers_items.append({
            'id': str(integration.id),
            'name': integration.name or integration.provider or str(integration.id),
            'provider': integration.provider,
            'balance': _to_float(amount),
            'debt': _to_float(debt),
            'netBalance': _to_float(net_amount),
            'currency': currency,
            'balanceUsd': _to_float(usd_amount),
            'balanceUpdatedAt': integration.balance_updated_at.isoformat() if integration.balance_updated_at else None,
            'debtUpdatedAt': getattr(integration, 'debt_updated_at', None).isoformat() if getattr(integration, 'debt_updated_at', None) else None,
        })

    providers_totals = [
        {
            'currency': code,
            'amount': _to_float(total_amount),
            'amountUsd': _to_float(convert_to_usd(total_amount, code)),
        }
        for code, total_amount in sorted(providers_totals_map.items())
    ]

    adjustments_items: List[dict] = []
    adjustments_totals_map: Dict[str, Decimal] = defaultdict(lambda: Decimal('0'))
    adjustments_total_usd = Decimal('0')
    adjustments_error: str | None = None
    try:
        adjustments_queryset = list(
            CapitalAdjustment.objects.filter(tenant_id=tenant_id).order_by('-created_at')
        )
    except ProgrammingError:
        adjustments_queryset = []
        adjustments_error = 'TABLE_MISSING'

    for adj in adjustments_queryset:
        currency = _normalize_currency_code(getattr(adj, 'currency', None))
        amount = _decimal_from(getattr(adj, 'amount', None))
        usd_amount = convert_to_usd(amount, currency)
        adjustments_total_usd += usd_amount
        adjustments_totals_map[currency] += amount
        adjustments_items.append({
            'id': str(adj.id),
            'label': getattr(adj, 'label', ''),
            'currency': currency,
            'amount': _to_float(amount),
            'amountUsd': _to_float(usd_amount),
            'note': getattr(adj, 'note', '') or '',
            'createdAt': adj.created_at.isoformat() if getattr(adj, 'created_at', None) else None,
            'updatedAt': adj.updated_at.isoformat() if getattr(adj, 'updated_at', None) else None,
        })
        overall_totals_map[currency] += amount
        currency_universe.add(currency)

    adjustments_totals = [
        {
            'currency': code,
            'amount': _to_float(total_amount),
            'amountUsd': _to_float(convert_to_usd(total_amount, code)),
        }
        for code, total_amount in sorted(adjustments_totals_map.items())
    ]

    # وسائـل الدفع الفعالة
    payment_methods_qs = PaymentMethod.objects.filter(tenant_id=tenant_id, is_active=True).order_by('name')
    payment_method_labels: List[str] = []
    seen_labels: set[str] = set()
    for pm in payment_methods_qs:
        label = _normalize_label_value(getattr(pm, 'name', ''))
        if label not in seen_labels:
            seen_labels.add(label)
            payment_method_labels.append(label)

    def _currency_sort_key(code: str) -> tuple[int, str]:
        if primary_currency and code == primary_currency:
            return (0, code)
        if code == 'USD':
            return (1, code)
        return (2, code)

    sorted_currencies = sorted(currency_universe, key=_currency_sort_key)

    overall_totals = [
        {
            'currency': code,
            'amount': _to_float(amount),
        }
        for code, amount in sorted(overall_totals_map.items(), key=lambda item: _currency_sort_key(item[0]))
    ]

    grand_total_usd = users_total_usd + providers_total_usd + adjustments_total_usd

    return {
        'rates': {code: _to_float(rate, places=6) for code, rate in rates_map.items()},
        'missingRates': sorted(code for code in missing_rates if code and code != 'USD'),
        'users': {
            'count': users_count,
            'totals': users_totals,
            'totalUsd': _to_float(users_total_usd),
        },
        'providers': {
            'items': providers_items,
            'totals': providers_totals,
            'totalUsd': _to_float(providers_total_usd),
        },
        'adjustments': {
            'items': adjustments_items,
            'totals': adjustments_totals,
            'totalUsd': _to_float(adjustments_total_usd),
            'error': adjustments_error,
        },
        'paymentMethods': payment_method_labels,
        'sortedCurrencies': sorted_currencies,
        'primaryCurrency': primary_currency,
        'overallTotals': overall_totals,
        'grandTotalUsd': _to_float(grand_total_usd),
    }

class AdminReportsCapitalView(APIView):
    permission_classes = [IsAuthenticated, RequireAdminRole]

    @extend_schema(tags=["Admin Reports"], responses={200: None})
    def get(self, request):
        tenant_id = _resolve_tenant_id(request)
        if not tenant_id:
            raise ValidationError('TENANT_ID_REQUIRED')
        payload = _build_capital_summary_payload(tenant_id)
        return Response(payload)


class AdminReportsCapitalExportView(APIView):
    permission_classes = [IsAuthenticated, RequireAdminRole]

    @extend_schema(tags=["Admin Reports"], responses={200: None})
    def get(self, request):
        tenant_id = _resolve_tenant_id(request)
        if not tenant_id:
            raise ValidationError('TENANT_ID_REQUIRED')

        summary = _build_capital_summary_payload(tenant_id)
        workbook = _build_capital_workbook(summary)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        timestamp = timezone.now().strftime('%Y%m%d_%H%M')
        filename = f'capital_report_{timestamp}.xlsx'

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class AdminReportsCapitalAdjustmentsView(APIView):
    permission_classes = [IsAuthenticated, RequireAdminRole]

    @extend_schema(tags=["Admin Reports"], request=CapitalAdjustmentInputSerializer, responses={201: CapitalAdjustmentSerializer})
    def post(self, request):
        tenant_id = _resolve_tenant_id(request)
        if not tenant_id:
            raise ValidationError('TENANT_ID_REQUIRED')
        payload = CapitalAdjustmentInputSerializer(data=request.data)
        payload.is_valid(raise_exception=True)
        data = payload.validated_data
        obj = CapitalAdjustment(
            tenant_id=tenant_id,
            label=data['label'],
            amount=data['amount'],
            currency=_normalize_currency_code(data.get('currency')),
            note=data.get('note') or '',
            created_by=request.user if getattr(request, 'user', None) and request.user.is_authenticated else None,
        )
        obj.save()
        return Response(CapitalAdjustmentSerializer(obj).data, status=201)


class AdminReportsCapitalAdjustmentDetailView(APIView):
    permission_classes = [IsAuthenticated, RequireAdminRole]

    def _get_object(self, tenant_id: str, adj_id: str) -> CapitalAdjustment:
        try:
            return CapitalAdjustment.objects.get(id=adj_id, tenant_id=tenant_id)
        except CapitalAdjustment.DoesNotExist:
            raise NotFound('CAPITAL_ADJUSTMENT_NOT_FOUND')

    @extend_schema(tags=["Admin Reports"], request=CapitalAdjustmentInputSerializer, responses={200: CapitalAdjustmentSerializer})
    def put(self, request, id: str):
        tenant_id = _resolve_tenant_id(request)
        if not tenant_id:
            raise ValidationError('TENANT_ID_REQUIRED')
        obj = self._get_object(tenant_id, id)
        payload = CapitalAdjustmentInputSerializer(data=request.data)
        payload.is_valid(raise_exception=True)
        data = payload.validated_data
        obj.label = data['label']
        obj.amount = data['amount']
        obj.currency = _normalize_currency_code(data.get('currency'))
        obj.note = data.get('note') or ''
        obj.save(update_fields=['label', 'amount', 'currency', 'note', 'updated_at'])
        return Response(CapitalAdjustmentSerializer(obj).data)

    @extend_schema(tags=["Admin Reports"], responses={204: None})
    def delete(self, request, id: str):
        tenant_id = _resolve_tenant_id(request)
        if not tenant_id:
            raise ValidationError('TENANT_ID_REQUIRED')
        obj = self._get_object(tenant_id, id)
        obj.delete()
        return Response(status=204)


class AdminReportsOverviewView(APIView):
    permission_classes = [IsAuthenticated]

    @extend_schema(
        tags=["Admin Reports"],
        parameters=[
            OpenApiParameter(name='X-Tenant-Host', required=False, type=str, location=OpenApiParameter.HEADER),
            OpenApiParameter(name='from', required=False, type=str, description='ISO date (YYYY-MM-DD)'),
            OpenApiParameter(name='to', required=False, type=str, description='ISO date (YYYY-MM-DD)'),
        ],
        responses={200: OverviewResponseSerializer}
    )
    def get(self, request):
        tenant_id = _resolve_tenant_id(request)
        if not tenant_id:
            raise ValidationError('TENANT_ID_REQUIRED')
        dfrom, dto = _parse_dates(request)
        # Orders: count + revenue per currency (sellPrice)
        oqs = ProductOrder.objects.filter(tenant_id=tenant_id, created_at__date__gte=dfrom, created_at__date__lte=dto)
        o_items = list(oqs.only('sell_price_currency','sell_price_amount'))
        orders = {
            'totalCount': oqs.count(),
            'totals': _money_totals(o_items, 'sell_price_currency', 'sell_price_amount'),
        }
        # Deposits
        dqs = Deposit.objects.filter(tenant_id=tenant_id, created_at__date__gte=dfrom, created_at__date__lte=dto)
        d_items = list(dqs.only('currency','amount'))
        deposits = {
            'totalCount': dqs.count(),
            'totals': _money_totals(d_items, 'currency', 'amount'),
        }
        # Payouts
        pqs = Payout.objects.filter(tenant_id=tenant_id, created_at__date__gte=dfrom, created_at__date__lte=dto)
        p_items = list(pqs.only('currency','amount'))
        payouts = {
            'totalCount': pqs.count(),
            'totals': _money_totals(p_items, 'currency', 'amount'),
        }
        return Response({ 'orders': orders, 'deposits': deposits, 'payouts': payouts })


class AdminReportsOverviewCSVView(APIView):
    permission_classes = [IsAuthenticated]

    @extend_schema(exclude=True)
    def get(self, request):
        tenant_id = _resolve_tenant_id(request)
        if not tenant_id:
            raise ValidationError('TENANT_ID_REQUIRED')
        dfrom, dto = _parse_dates(request)
        # Reuse calculations similar to JSON endpoint
        oqs = ProductOrder.objects.filter(tenant_id=tenant_id, created_at__date__gte=dfrom, created_at__date__lte=dto)
        o_items = list(oqs.only('sell_price_currency','sell_price_amount'))
        o_totals = _money_totals(o_items, 'sell_price_currency', 'sell_price_amount')
        dqs = Deposit.objects.filter(tenant_id=tenant_id, created_at__date__gte=dfrom, created_at__date__lte=dto)
        d_items = list(dqs.only('currency','amount'))
        d_totals = _money_totals(d_items, 'currency', 'amount')
        pqs = Payout.objects.filter(tenant_id=tenant_id, created_at__date__gte=dfrom, created_at__date__lte=dto)
        p_items = list(pqs.only('currency','amount'))
        p_totals = _money_totals(p_items, 'currency', 'amount')

        resp = HttpResponse(content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = 'attachment; filename="reports_overview.csv"'
        writer = csv.writer(resp)
        writer.writerow(["section","totalCount","currency","amount","from","to"]) 
        for t in o_totals:
            writer.writerow(["orders", oqs.count(), t['currency'], t['amount'], dfrom.isoformat(), dto.isoformat()])
        for t in d_totals:
            writer.writerow(["deposits", dqs.count(), t['currency'], t['amount'], dfrom.isoformat(), dto.isoformat()])
        for t in p_totals:
            writer.writerow(["payouts", pqs.count(), t['currency'], t['amount'], dfrom.isoformat(), dto.isoformat()])
        return resp


class _DailyBase(APIView):
    permission_classes = [IsAuthenticated]
    currency_field: str = 'currency'
    amount_field: str = 'amount'
    tags: list[str] = ["Admin Reports"]

    def _get_qs(self, tenant_id, dfrom, dto):
        raise NotImplementedError

    @extend_schema(
        parameters=[
            OpenApiParameter(name='X-Tenant-Host', required=False, type=str, location=OpenApiParameter.HEADER),
            OpenApiParameter(name='from', required=False, type=str, description='ISO date (YYYY-MM-DD)'),
            OpenApiParameter(name='to', required=False, type=str, description='ISO date (YYYY-MM-DD)'),
        ],
        responses={200: DailyResponseSerializer}
    )
    def get(self, request):
        tenant_id = _resolve_tenant_id(request)
        if not tenant_id:
            raise ValidationError('TENANT_ID_REQUIRED')
        dfrom, dto = _parse_dates(request)
        qs = self._get_qs(tenant_id, dfrom, dto)
        # Group by date
        agg = qs.annotate(d=TruncDate('created_at')).values('d', self.currency_field).annotate(count=Count('id'), total=Sum(self.amount_field)).order_by('d')
        # Collect per-day totals per currency
        by_day = {}
        for row in agg:
            dt = row['d']
            cur = row[self.currency_field] or 'USD'
            total = float(row['total'] or 0)
            if dt not in by_day:
                by_day[dt] = { 'date': dt, 'count': 0, 'totals': {} }
            by_day[dt]['count'] += int(row['count'] or 0)
            by_day[dt]['totals'][cur] = round(by_day[dt]['totals'].get(cur, 0) + total, 6)
        # Transform
        items = []
        for dt in sorted(by_day.keys()):
            totals = [{ 'currency': c, 'amount': amt } for c, amt in sorted(by_day[dt]['totals'].items())]
            items.append({ 'date': dt, 'count': by_day[dt]['count'], 'totals': totals })
        return Response({ 'items': items })


class _DailyCSVBase(_DailyBase):
    @extend_schema(exclude=True)
    def get(self, request):
        tenant_id = _resolve_tenant_id(request)
        if not tenant_id:
            raise ValidationError('TENANT_ID_REQUIRED')
        dfrom, dto = _parse_dates(request)
        qs = self._get_qs(tenant_id, dfrom, dto)
        agg = qs.annotate(d=TruncDate('created_at')).values('d', self.currency_field).annotate(count=Sum('id') * 0 + Count('id'), total=Sum(self.amount_field)).order_by('d')
        # Prepare CSV
        resp = HttpResponse(content_type='text/csv; charset=utf-8')
        fname = getattr(self, 'filename', 'daily.csv')
        resp['Content-Disposition'] = f'attachment; filename="{fname}"'
        writer = csv.writer(resp)
        writer.writerow(["date","currency","amount","count","from","to"]) 
        for row in agg:
            dt = row['d']
            cur = row.get(self.currency_field) or 'USD'
            total = float(row['total'] or 0)
            cnt = int(row['count'] or 0)
            writer.writerow([dt.isoformat() if hasattr(dt,'isoformat') else str(dt), cur, round(total,6), cnt, dfrom.isoformat(), dto.isoformat()])
        return resp


class AdminOrdersDailyView(_DailyBase):
    currency_field = 'sell_price_currency'
    amount_field = 'sell_price_amount'

    @extend_schema(tags=["Admin Reports"])
    def get(self, request):
        return super().get(request)

    def _get_qs(self, tenant_id, dfrom, dto):
        return ProductOrder.objects.filter(tenant_id=tenant_id, created_at__date__gte=dfrom, created_at__date__lte=dto)


class AdminOrdersDailyCSVView(_DailyCSVBase):
    currency_field = 'sell_price_currency'
    amount_field = 'sell_price_amount'
    filename = 'orders_daily.csv'

    def _get_qs(self, tenant_id, dfrom, dto):
        return ProductOrder.objects.filter(tenant_id=tenant_id, created_at__date__gte=dfrom, created_at__date__lte=dto)


class AdminDepositsDailyView(_DailyBase):
    @extend_schema(tags=["Admin Reports"])
    def get(self, request):
        return super().get(request)

    def _get_qs(self, tenant_id, dfrom, dto):
        return Deposit.objects.filter(tenant_id=tenant_id, created_at__date__gte=dfrom, created_at__date__lte=dto)


class AdminDepositsDailyCSVView(_DailyCSVBase):
    filename = 'deposits_daily.csv'

    def _get_qs(self, tenant_id, dfrom, dto):
        return Deposit.objects.filter(tenant_id=tenant_id, created_at__date__gte=dfrom, created_at__date__lte=dto)


class AdminPayoutsDailyView(_DailyBase):
    @extend_schema(tags=["Admin Reports"])
    def get(self, request):
        return super().get(request)

    def _get_qs(self, tenant_id, dfrom, dto):
        return Payout.objects.filter(tenant_id=tenant_id, created_at__date__gte=dfrom, created_at__date__lte=dto)


class AdminPayoutsDailyCSVView(_DailyCSVBase):
    filename = 'payouts_daily.csv'

    def _get_qs(self, tenant_id, dfrom, dto):
        return Payout.objects.filter(tenant_id=tenant_id, created_at__date__gte=dfrom, created_at__date__lte=dto)
